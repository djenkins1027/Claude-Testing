import io

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def read_docx(content: bytes) -> str:
    doc = Document(io.BytesIO(content))
    lines = []
    for p in doc.paragraphs:
        style = p.style.name if p.style else "Normal"
        if style == "Heading 1":
            lines.append(f"# {p.text}")
        elif style == "Heading 2":
            lines.append(f"## {p.text}")
        elif style == "List Bullet":
            lines.append(f"- {p.text}")
        else:
            lines.append(p.text)
    return "\n".join(lines)


def read_xlsx(content: bytes) -> list[list]:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def read_pdf(content: bytes) -> str:
    reader = PdfReader(io.BytesIO(content))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


READERS = {"docx": read_docx, "xlsx": read_xlsx, "pdf": read_pdf}
